"TimeSheet Project"

import tkinter as tk
import openpyxl as ox
import getpass
import os
from datetime import datetime, date, time, timezone, timedelta

###################
# TimeSheet Class #
###################

class TimeSheet:

    def __init__(self, studentName, dollarsPerHour):

        self.name = studentName
        self.dollars = dollarsPerHour

        self.book = ox.Workbook()
        
        self.book.create_sheet("Lesson", 0)
        self.book.create_sheet("Practice", 1)

        self.filename = str(self.name + ".xlsx")

        self.book.save(filename = self.filename) #'/Users/'+getpass.getuser()+'/Desktop/Personal_Projects/' + self.filename


    def startLesson(self):
        """Iterates through the rows in Column A until it finds an empty row; saves the empty row number to newRow"""
        pathToFile = str(os.path.abspath(self.filename)) #'/Users/'+getpass.getuser()+'/Desktop/Personal_Projects/' + self.filename 
        self.book = ox.load_workbook(pathToFile) 

        for row in self.book["Lesson"].iter_rows(min_row = 1, max_row = 1000, min_col = None, max_col = 5, values_only = False):
            if row[0].value == None:
                newRow = row
                break
        
        newRow[0].value = datetime.now().date()
        newRow[1].value = datetime.now().time()

        self.book.save(self.filename)

    def startPractice(self):
        """Iterates through the rows in Column A until it finds an empty row; saves the empty row number to newRow"""

        pathToFile = str(os.path.abspath(self.filename)) #'/Users/'+getpass.getuser()+'/Desktop/Personal_Projects/' + self.filename 
        self.book = ox.load_workbook(pathToFile)

        for row in self.book["Practice"].iter_rows(min_row = 1, max_row = 1000, min_col = None, max_col = 5, values_only = False):
            if row[0].value == None:
                newRow = row
                break
        
        newRow[0].value = datetime.now().date()
        newRow[1].value = datetime.now().time()
        
        self.book.save(self.filename)

    def stopLesson(self):
        "Iterates through the rows until it finds a row with Column A filled and Column C empty; saves row to endRow"

        pathToFile = str(os.path.abspath(self.filename)) #'/Users/'+getpass.getuser()+'/Desktop/Personal_Projects/' + self.filename 
        self.book = ox.load_workbook(pathToFile)
        
        for row in self.book["Lesson"].iter_rows(min_row = 1, max_row = 1000, min_col = None, max_col = 5, values_only = False):
            if row[0].value != None and row[3].value == None:
                endRow = row
                break
        #else:
            #return None
        
        endRow[2].value = datetime.now().time()
        start = datetime.combine(endRow[0].value, endRow[2].value)
        end = datetime.combine(endRow[0].value, endRow[1].value)
        timeDif = start - end
        endRow[3].value = timeDif #datetime.combine(endRow[0].value, endRow[2].value) - datetime.combine(endRow[0].value, endRow[1].value)
        endRow[4].value = timeDif.total_seconds()
        self.book.save(self.filename)

    def stopPractice(self):
        "Iterates through the rows until it finds a row with Column A filled and Column C empty; saves row to endRow"

        pathToFile = str(os.path.abspath(self.filename)) #'/Users/'+getpass.getuser()+'/Desktop/Personal_Projects/' + self.filename 
        self.book = ox.load_workbook(pathToFile)

        for row in self.book["Practice"].iter_rows(min_row = 1, max_row = 1000, min_col = None, max_col = 5, values_only = False):
            if row[0].value != None and row[3].value == None:
                endRow = row
                break
        
        endRow[2].value = datetime.now().time()
        start = datetime.combine(endRow[0].value, endRow[2].value)
        end = datetime.combine(endRow[0].value, endRow[1].value)
        timeDif = start - end
        endRow[3].value = timeDif #datetime.combine(endRow[0].value, endRow[2].value) - datetime.combine(endRow[0].value, endRow[1].value)
        endRow[4].value = timeDif.total_seconds()
        self.book.save(self.filename)

    def earnings(self):
        """Sums the numbers in Column D for both Lesson and Practice sheets; variable lTime is the total time in Lesson; 
        pTime is the total time in Practice; displays min(lTime, pTime) * dollarsPerHour"""

        pathToFile = str(os.path.abspath(self.filename)) #'/Users/'+getpass.getuser()+'/Desktop/Personal_Projects/' + self.filename 
        self.book = ox.load_workbook(pathToFile)

        lTime, pTime = 0, 0
        for cell in self.book["Lesson"]['E']:
            if cell.value == None:
                break
            lTime += cell.value

        for cell in self.book["Practice"]['E']:
            if cell.value == None:
                break
            pTime += cell.value
        
        delta = min(lTime, pTime)
        pay = (delta // (60 * 60)) * self.dollars
    
        return pay 
    
#######
# GUI #
#######

root = tk.Tk()
root.title("TimeSheet")

#BF = button function
buttonsDict = {}
studentsDict = {}
locationsDict = {}
studentRow = 1
def newStudentBF(name):
    global person
    global studentRow

    person = TimeSheet(name, dollarsPerHour=10)
    intermediateFunc = lambda buttonName: lambda: pickStudent(buttonName)
    studentButton = tk.Button(root, text = person.name, command = intermediateFunc(name)) 
    studentButton.grid(row = studentRow, column = 0, pady = 20, padx = 10)
    buttonsDict[name] = studentButton
    locationsDict[studentButton] = studentButton.grid_info()
    studentsDict[studentButton] = person

    studentRow += 1

    newInput.grid_forget()
    newCreateButton.grid_forget()

def pickStudent(name):
    #this function should set global variable student to the specific student object which was selected
    global globalStudent
    global startLButton
    global startPButton
    global stopLButton
    global stopPButton
    global saveReturnButton
    global earningsButton
    global studentLabel
    global earningsLabel
    
    for name in buttonsDict:
        buttonsDict[name].grid_forget()
    newButton.grid_forget()

    globalStudent = studentsDict[buttonsDict[name]]
    startLButton = tk.Button(root, text = "Start Lesson", pady = 20, padx = 10, command = lambda: startLessonBF(globalStudent))
    startPButton = tk.Button(root, text = "Start Practice", pady = 20, padx = 10, command = lambda: startPracticeBF(globalStudent))
    stopLButton = tk.Button(root, text = "Stop Lesson", pady = 20, padx = 10, command = lambda: stopLessonBF(globalStudent))
    stopPButton = tk.Button(root, text = "Stop Practice", pady = 20, padx = 10, command = lambda: stopPracticeBF(globalStudent))
    saveReturnButton = tk.Button(root, text = "Back", command = backBF)
    earningsButton = tk.Button(root, text = "Display Earnings", pady = 20, padx = 10, command = lambda: earningsBF(globalStudent))
    studentLabel = tk.Label(root, text = name)
    earningsLabel = tk.Label(root)

    studentLabel.grid(row = 0, column = 2)
    startLButton.grid(row = 1, column = 1)
    startPButton.grid(row = 1, column = 3)
    earningsButton.grid(row = 3, column = 2)
    saveReturnButton.grid(row = 0, column = 1)

def backBF():
    studentLabel.grid_forget()
    startLButton.grid_forget()
    startPButton.grid_forget()
    stopLButton.grid_forget()
    stopPButton.grid_forget()
    earningsButton.grid_forget()
    earningsLabel.grid_forget()
    saveReturnButton.grid_forget()

    newButton.grid(row = 0, column = 0, pady = 20, padx = 10)
    for name in buttonsDict:
        buttonsDict[name].grid(row = locationsDict[buttonsDict[name]]["row"], column = locationsDict[buttonsDict[name]]["column"], pady = 20, padx = 10)

def startLessonBF(student):
    student.startLesson()
    startLButton.grid_forget()
    stopLButton.grid(row = 1, column = 1)

def startPracticeBF(student):
    student.startPractice()
    startPButton.grid_forget()
    stopPButton.grid(row = 1, column = 3)

def stopLessonBF(student):
    student.stopLesson()
    stopLButton.grid_forget()
    startLButton.grid(row = 1, column = 1)

def stopPracticeBF(student):
    student.stopPractice()
    stopPButton.grid_forget()
    startPButton.grid(row = 1, column = 3)

def earningsBF(student):
    global earningsLabel
    payment = student.earnings()
    earningsLabel.grid_forget()
    earningsLabel = tk.Label(root, text = "Dollars Earned: " + str(payment))
    earningsLabel.grid(row = 4, column = 2)

def newStudentInputBF():
    global newInput 
    global newCreateButton

    newInput = tk.Entry(root)
    newInput.grid(row = 0, column = 1)

    newCreateButton = tk.Button(root, text = "Create", command = lambda: newStudentBF(newInput.get()))
    newCreateButton.grid(row = 1, column = 1)

newButton = tk.Button(root, text = "New Student", command = newStudentInputBF)
newButton.grid(row = 0, column = 0, pady = 20, padx = 10)

root.mainloop()


